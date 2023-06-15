import pandas as pd

from openpyxl import load_workbook

from openpyxl.chart import BarChart, Reference

from openpyxl.styles import Font


# from pathlib import Path
# import zipfile

# Renombrar nombre de carpeta
#path = Path('test')
#path.rename('nuevo_test')

# Renombrar archivo dentro de una carpeta
#path = Path('nuevo_test/gastos.txt')
#nuevoNombrePath = path.with_name('gastos-diciembre.txt')
#path.rename(nuevoNombrePath)

#Obtener path de los subdirectorios inmediatos
# carpeta = Path('2023')
# for path in list(carpeta.iterdir()):
#     print(path)

# Obtener path de todos los subdirectorios
# carpeta = Path('./nuevo_test')
#
# # paths = carpeta.glob('**/*')
# # for path in paths:
# #     print(path)
# paths = carpeta.glob('**/*')
#
# for path in paths:
#      if path.is_file():
#          print(path)
#      else:
#          print('No hay archivo')


# folder = Path('extensiones')

# Pasar de txt a csv
# for path in list(folder.iterdir()):
#     print(path)
#     if path.suffix == '.txt':
#         nuevoNombreExtension = path.with_suffix('.csv')
#         path.rename(nuevoNombreExtension)

# Pasar de csv a txt
# for path in folder.glob('**/*.csv'):
#     print(path)
#     nuevoNombreExtension = path.with_suffix('.txt')
#     path.rename((nuevoNombreExtensinumeros

# numeros = [1,2,3,4,5,6,7,8,9]

# crear archivos txt
# for i in numeros:
#     with open(f'test{i}.txt', 'w') as file:
#         file.write('Hola Mundo')


# # Eliminar todos los archivos txt
# for path in Path('.').glob('*.txt'):
#     print(path)
#     path.unlink()


# Eliminar todos los archivos txt menos el test 9
# for path in Path('.').glob('test[1-8].txt'):
#     print(path)
#     path.unlink()

# Crea una carpeta temp y extrae un archivo .zip en esa carpeta
# directorio_actual = Path('.')
# directorio_objetivo = Path('temp')
#
# for path in directorio_actual.glob('*.zip'):
#     print(path)
#     with zipfile.ZipFile(path, 'r') as zipObj:
#         zipObj.extractall(path=directorio_objetivo)

# simp = pd.read_html('https://en.wikipedia.org/wiki/List_of_The_Simpsons_episodes_(seasons_1%E2%80%9320)')
# print(simp[4])


archivo_excel = pd.read_excel('excel/supermarket_sales.xlsx')
# print(archivo_excel[['Gender', 'Product line', 'Total']])

table_pivote = archivo_excel.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
print(table_pivote)

table_pivote.to_excel('excel/sales_2022.xlsx', startrow=4, sheet_name='Report')

wb =load_workbook('excel/sales_2022.xlsx')
pestania = wb['Report']

min_col = wb.active.min_column
max_col = wb.active.max_column
min_fila = wb.active.min_row
max_fila = wb.active.max_row

print(min_col)
print(max_col)
print(min_fila)
print(max_col)

# Graficos
barchart = BarChart()

data = Reference(pestania, min_col= min_col+1, max_col=max_col, min_row=min_fila, max_row=max_fila)
categoria = Reference(pestania, min_col= min_col, max_col=min_col, min_row=min_fila+1, max_row=max_fila)
barchart.add_data(data, titles_from_data=True)
barchart.add_data(categoria)
pestania.add_chart(barchart,'B12')
barchart.title ='Ventas'
barchart.style = 3

pestania['A1'] = 'Reporte'
pestania['A2'] = '2022'
pestania['A1'].font = Font('Arial', bold=True, size=20)
pestania['A2'].font = Font('Arial', bold=True, size=15)

wb.save('./excel/sales_2022.xlsx')